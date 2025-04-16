from epyt import epanet
import numpy as np
d = epanet('FILEGOCCUADUY.inp') 
import pandas as pd
import openpyxl

# # Định nghĩa các cột dựa trên yêu cầu của d.getNodeBasedemand
# columns = ['NodeID', 'Demand', 'Pattern', 'Category']

# # Lấy danh sách các tên NodeID từ lệnh d.getNodeNameID
# node_ids = d.getNodeNameID()

# # Tạo một DataFrame rỗng với các cột đã chỉ định
# df = pd.DataFrame(columns=columns)

# # Thêm các tên NodeID vào cột 'NodeID' của DataFrame
# df['NodeID'] = node_ids

# # Lưu DataFrame vào một file Excel
# excel_file_path = r'C:\Users\Admin\Documents\VSCODE\DATA.xlsx'
# df.to_excel(excel_file_path, index=False)

########################################################################################################
#Phần cập nhật chỉ số Demand
# Mở file Excel
workbook = openpyxl.load_workbook(r'C:\Users\Admin\Documents\VSCODE\DATA.xlsx')

# Lấy sheet hiện tại (active sheet) trong workbook
sheet = workbook.active

# Lấy danh sách các chỉ số Node từ lệnh d.getNodeIndex
node_indices = d.getNodeIndex()

# Chạy lệnh d.addNodeJunctionDemand cho từng hàng từ 2 đến 123
for row in range(2, 124):
    node_id = str(node_indices[row - 2])  # Lấy NodeID từ danh sách node_indices
    demand = float(sheet.cell(row=row, column=2).value)
    pattern = str(sheet.cell(row=row, column=3).value)
    category = str(sheet.cell(row=row, column=4).value)
    if node_id and demand and pattern and category:
        d.addNodeJunctionDemand(node_id, demand, pattern, category)

print(d.getNodeBaseDemands(1))